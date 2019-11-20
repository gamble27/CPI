from openpyxl import load_workbook


def make_data_numeric(full_file_name):
    wb = load_workbook(filename=full_file_name)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.value = str(cell.value).replace(',', '.')


# actually, this one did not work out
if __name__ == "__main__":
    files = [f"dataset/ukrstat_dno/bisc_201{i}_e.xlsx" for i in range(6, 9)]

    for file in files:
        make_data_numeric(file)
